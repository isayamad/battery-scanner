"""
Battery Scanner App - Simple Working Version
Logs battery IDs with timestamp, prevents duplicates
"""

import os
from datetime import datetime
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available")

class BatteryScannerApp(App):
    def build(self):
        # Set up the main layout
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Title
        title = Label(text="🔋 Battery Scanner", font_size=24, size_hint_y=0.1)
        layout.add_widget(title)
        
        # Status label
        self.status_label = Label(text="Ready to scan", size_hint_y=0.1)
        layout.add_widget(self.status_label)
        
        # Manual entry button (since camera scanning requires extra setup)
        manual_btn = Button(text="📝 Enter Battery ID", size_hint_y=0.15, font_size=18)
        manual_btn.bind(on_press=self.open_manual_entry)
        layout.add_widget(manual_btn)
        
        # View log button
        view_btn = Button(text="📋 View Log", size_hint_y=0.15, font_size=18)
        view_btn.bind(on_press=self.view_log)
        layout.add_widget(view_btn)
        
        # Count display
        self.count_label = Label(text="Total logged: 0", size_hint_y=0.1)
        layout.add_widget(self.count_label)
        
        # Initialize Excel file
        self.excel_path = os.path.join(os.path.expanduser("~"), "BatteryLog.xlsx")
        self.init_excel()
        self.update_count()
        
        return layout
    
    def init_excel(self):
        """Create Excel file if it doesn't exist"""
        if not EXCEL_AVAILABLE:
            return
        
        if not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Battery Log"
            ws['A1'] = "Battery ID"
            ws['B1'] = "Date"
            ws['C1'] = "Time"
            
            # Style the header
            for col in ['A', 'B', 'C']:
                ws[f'{col}1'].font = Font(bold=True, color="FFFFFF")
                ws[f'{col}1'].fill = PatternFill("solid", fgColor="1F78B4")
                ws.column_dimensions[col].width = 20
            
            wb.save(self.excel_path)
    
    def get_total_count(self):
        """Get total number of logged batteries"""
        if not EXCEL_AVAILABLE or not os.path.exists(self.excel_path):
            return 0
        
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        count = max(0, wb.active.max_row - 1)
        wb.close()
        return count
    
    def update_count(self):
        """Update the count display"""
        count = self.get_total_count()
        self.count_label.text = f"Total logged: {count}"
    
    def open_manual_entry(self, instance):
        """Open popup for manual ID entry"""
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        content.add_widget(Label(text="Enter Battery ID:", size_hint_y=0.2))
        
        text_input = TextInput(
            hint_text="e.g. LEADPOWER 25041816135",
            multiline=False,
            font_size=16,
            size_hint_y=0.3
        )
        content.add_widget(text_input)
        
        button_box = BoxLayout(orientation='horizontal', spacing=10, size_hint_y=0.2)
        submit_btn = Button(text="Submit")
        cancel_btn = Button(text="Cancel")
        button_box.add_widget(submit_btn)
        button_box.add_widget(cancel_btn)
        content.add_widget(button_box)
        
        popup = Popup(
            title="Manual Entry",
            content=content,
            size_hint=(0.9, 0.4)
        )
        
        def submit_action(btn):
            battery_id = text_input.text.strip()
            if battery_id:
                self.log_battery(battery_id)
                popup.dismiss()
        
        submit_btn.bind(on_press=submit_action)
        cancel_btn.bind(on_press=popup.dismiss)
        popup.open()
    
    def log_battery(self, battery_id):
        """Add battery to log if not duplicate"""
        if not EXCEL_AVAILABLE:
            self.status_label.text = "Excel support not available"
            return
        
        # Check if file exists, create if not
        if not os.path.exists(self.excel_path):
            self.init_excel()
        
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        
        # Check for duplicate
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == battery_id:
                self.status_label.text = f"⚠️ DUPLICATE: {battery_id} already logged on {row[1]}"
                wb.close()
                return
        
        # Add new entry
        now = datetime.now()
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=battery_id)
        ws.cell(row=next_row, column=2, value=now.strftime("%Y-%m-%d"))
        ws.cell(row=next_row, column=3, value=now.strftime("%H:%M:%S"))
        
        wb.save(self.excel_path)
        
        self.status_label.text = f"✅ Added: {battery_id}"
        self.update_count()
    
    def view_log(self, instance):
        """Show all logged batteries"""
        if not EXCEL_AVAILABLE or not os.path.exists(self.excel_path):
            self.status_label.text = "No log file found"
            return
        
        # Create scrollable content
        scroll = ScrollView()
        log_layout = BoxLayout(orientation='vertical', size_hint_y=None, spacing=2)
        log_layout.bind(minimum_height=log_layout.setter('height'))
        
        # Read all entries
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        entries = []
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row[0]:
                entries.append(f"{row[0]} - {row[1]} {row[2]}")
        wb.close()
        
        if not entries:
            log_layout.add_widget(Label(text="No entries yet", size_hint_y=None, height=40))
        else:
            # Show last 50 entries (newest first)
            for entry in reversed(entries[-50:]):
                label = Label(
                    text=entry,
                    size_hint_y=None,
                    height=35,
                    font_size=12,
                    text_size=(300, None),
                    halign='left'
                )
                log_layout.add_widget(label)
        
        scroll.add_widget(log_layout)
        
        popup = Popup(
            title="Battery Log",
            content=scroll,
            size_hint=(0.95, 0.8)
        )
        popup.open()

if __name__ == "__main__":
    BatteryScannerApp().run()
